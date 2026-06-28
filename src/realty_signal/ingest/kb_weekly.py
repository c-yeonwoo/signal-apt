"""KB 주간 시계열(주택가격동향) 엑셀 파서.

KB부동산이 매주 공개하는 '주간 시계열' 엑셀을 tidy long DataFrame 으로 변환한다.
대상 시트:
  - 매수매도   → 매수우위지수 (buyer_superiority)   0~200, >100 매수자 우세
  - 전세수급   → 전세수급지수 (jeonse_supply)        0~200, >100 수요 우세
  - 매매증감   → 매매가격 증감률 (sale_change, %)
  - 전세증감   → 전세가격 증감률 (jeonse_change, %)

모든 시트는 동일한 주(week) 행을 공유한다(5행=첫 주). 날짜는 datetime 셀이
가장 많은 시트를 기준 달력으로 삼아 행 인덱스로 정렬한다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DATA_START_ROW = 5  # 실데이터 시작 행 (1-indexed)

# 시트명 → {서브헤더 라벨: metric 키}. 3열 그룹(우위/열위/지수)에서 필요한 열만 추출.
#   buyer_demand(매수세우위 raw)는 "5/10/15/20" 사다리용,
#   buyer_superiority(매수우위지수)는 차트용 — 별개 지표로 다룬다.
GROUPED_SHEETS = {
    "매수매도": {"매수세우위": "buyer_demand", "매수우위지수": "buyer_superiority"},
    "전세수급": {"전세수급지수": "jeonse_supply"},
}
# (시트명, metric 키) — 지역당 1열, 헤더행은 '전국'으로 자동 탐지
CHANGE_SHEETS = {
    "매매증감": "sale_change",
    "전세증감": "jeonse_change",
}

_ASCII = re.compile(r"[A-Za-z]")


def _clean_region(raw: object) -> str:
    """'전국 Total' → '전국', '6개광역시 6 Large Cities' → '6개광역시'."""
    s = str(raw).replace("\n", " ").strip()
    m = _ASCII.search(s)
    if m:
        s = s[: m.start()]
    # 영문 번역부가 숫자로 시작하는 경우('6 Large Cities') 꼬리 숫자/공백 제거
    return s.rstrip("0123456789 ").strip()


def _normalize_dates(raw: list[object]) -> list[pd.Timestamp]:
    """KB 의 축약 날짜('08.4.14, 4.21, 5.5 ...)를 연속 Timestamp 로 복원."""
    out: list[pd.Timestamp] = []
    year: int | None = None
    prev_month: int | None = None
    for v in raw:
        if isinstance(v, datetime):
            year, prev_month = v.year, v.month
            out.append(pd.Timestamp(v))
            continue
        if v is None or str(v).strip() == "":
            out.append(pd.NaT)
            continue
        parts = [p for p in str(v).strip().lstrip("'").split(".") if p != ""]
        try:
            if len(parts) == 3:
                yy, mm, dd = (int(p) for p in parts)
                year = 2000 + yy if yy < 100 else yy
                month, day = mm, dd
            elif len(parts) == 2:
                month, day = int(parts[0]), int(parts[1])
                if prev_month is not None and month < prev_month:
                    year = (year or 2008) + 1
            else:
                out.append(pd.NaT)
                continue
            prev_month = month
            out.append(pd.Timestamp(year=year, month=month, day=day))
        except (ValueError, TypeError):
            out.append(pd.NaT)
    return out


@dataclass
class KBWeekly:
    """파싱된 KB 주간 시계열. `long` 은 (date, region, metric, value) tidy frame."""

    long: pd.DataFrame

    @property
    def regions(self) -> list[str]:
        return sorted(self.long["region"].unique())

    @property
    def metrics(self) -> list[str]:
        return sorted(self.long["metric"].unique())

    @property
    def last_date(self) -> pd.Timestamp:
        return self.long["date"].max()

    def latest(self) -> pd.DataFrame:
        """지역 × metric 최근값 피벗 (가장 최근 non-null 주)."""
        df = self.long.dropna(subset=["value"]).sort_values("date")
        latest = df.groupby(["region", "metric"]).tail(1)
        return latest.pivot(index="region", columns="metric", values="value")

    def series(self, region: str, metric: str) -> pd.Series:
        """특정 지역·metric 의 주간 시계열."""
        m = (self.long["region"] == region) & (self.long["metric"] == metric)
        s = self.long[m].set_index("date")["value"].dropna().sort_index()
        s.name = f"{region}/{metric}"
        return s


def _parse_grouped(ws, label_map: dict[str, str], dates: list[pd.Timestamp]) -> pd.DataFrame:
    """3열 그룹(우위/열위/지수) 시트에서 label_map 에 지정한 서브헤더 열들을 long 으로."""
    region_at: dict[int, str] = {
        c: _clean_region(ws.cell(row=2, column=c).value)
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=2, column=c).value
    }
    cols = sorted(region_at)
    rows = []
    for c in range(1, ws.max_column + 1):
        metric = label_map.get(str(ws.cell(row=3, column=c).value).strip())
        if metric is None:
            continue
        region = next((region_at[rc] for rc in reversed(cols) if rc <= c), None)
        if not region:
            continue
        for i, r in enumerate(range(DATA_START_ROW, DATA_START_ROW + len(dates))):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                rows.append((dates[i], region, metric, float(v)))
    return pd.DataFrame(rows, columns=["date", "region", "metric", "value"])


def _find_header_row(ws) -> int:
    for r in range(1, 6):
        for c in range(1, min(ws.max_column, 10) + 1):
            if str(ws.cell(row=r, column=c).value).strip() == "전국":
                return r
    raise ValueError(f"'전국' 헤더를 {ws.title} 에서 찾지 못함")


def _parse_change(ws, metric: str, dates: list[pd.Timestamp]) -> pd.DataFrame:
    """지역당 1열인 증감률 시트 파싱."""
    hr = _find_header_row(ws)
    region_at = {
        c: str(ws.cell(row=hr, column=c).value).strip()
        for c in range(2, ws.max_column + 1)
        if ws.cell(row=hr, column=c).value
    }
    rows = []
    for c, region in region_at.items():
        for i, r in enumerate(range(DATA_START_ROW, DATA_START_ROW + len(dates))):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                rows.append((dates[i], region, metric, float(v)))
    return pd.DataFrame(rows, columns=["date", "region", "metric", "value"])


def _read_dates(ws, n_rows: int) -> list[pd.Timestamp]:
    raw = [ws.cell(row=r, column=1).value for r in range(DATA_START_ROW, DATA_START_ROW + n_rows)]
    return _normalize_dates(raw)


def load(path: str | Path) -> KBWeekly:
    """KB 주간 시계열 엑셀을 읽어 KBWeekly 반환."""
    # read_only 모드는 무작위 .cell() 접근이 극도로 느려 일반 모드로 로드한다.
    wb = load_workbook(path, data_only=True)

    # 기준 달력: datetime 셀이 가장 많은 시트의 A열
    best_dates: list[pd.Timestamp] = []
    n_rows = 0
    for sh in list(GROUPED_SHEETS) + list(CHANGE_SHEETS):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        rows = ws.max_row - DATA_START_ROW + 1
        dates = _read_dates(ws, rows)
        valid = sum(d is not pd.NaT and pd.notna(d) for d in dates)
        if valid > sum(d is not pd.NaT and pd.notna(d) for d in best_dates):
            best_dates, n_rows = dates, rows

    frames = []
    for sh, label_map in GROUPED_SHEETS.items():
        if sh in wb.sheetnames:
            frames.append(_parse_grouped(wb[sh], label_map, best_dates[:n_rows]))
    for sh, metric in CHANGE_SHEETS.items():
        if sh in wb.sheetnames:
            frames.append(_parse_change(wb[sh], metric, best_dates[:n_rows]))

    long = pd.concat(frames, ignore_index=True).dropna(subset=["date"])
    return KBWeekly(long=long)
